import csv
import os
from openpyxl import load_workbook, Workbook


class CsvManager:

    def __init__(self):
        self.path = None
        self.sheets = {}  # {name: list[list]}

    # -------------------------
    # load
    # -------------------------

    def load(self, path):
        self.path = path
        ext = os.path.splitext(path)[1].lower()

        if ext == ".csv":
            self._load_csv(path)
        else:
            self._load_excel(path)

    def _load_csv(self, path):
        with open(path, newline="", encoding="utf-8") as f:
            self.sheets = {"Sheet1": list(csv.reader(f))}

    def _load_excel(self, path):
        wb = load_workbook(path, data_only=True)
        self.sheets = {}

        for ws in wb:
            self.sheets[ws.title] = [
                list(r) for r in ws.iter_rows(values_only=True)
            ]

    # -------------------------
    # save
    # -------------------------

    def save(self, path=None):
        path = path or self.path
        ext = os.path.splitext(path)[1].lower()

        if ext == ".csv":
            self._save_csv(path)
        else:
            self._save_excel(path)

    def _save_csv(self, path):
        active_view = self.sheet_tabs.get_active_view()
        if active_view:
            # Get the sheet name from the SheetTabs widget
            active_sheet_name = self.sheet_tabs.tab(active_view.master, "text")
    
            with open(path, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerows(self.sheets[active_sheet_name])
        else:
            print("No active sheet found.")  # Handle the case where there's no active sheet



    def _save_excel(self, path):
        wb = Workbook()

        first = True
        for name, data in self.sheets.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = name
            first = False

            for row in data:
                ws.append(row)

        wb.save(path)


    def _new_sheet(self, name=None):
        if name is None:
            base = "Sheet"
            i = 1
            while f"{base}{i}" in self.sheets:
                i += 1
            name = f"{base}{i}"
    
        self.sheets[name] = [[""]]
        return name

