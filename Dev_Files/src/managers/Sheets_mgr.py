# sheets_mgr.py
import csv
from openpyxl import load_workbook, Workbook


class SheetsManager:
    def __init__(self):
        self.sheets = {}  # {name: list[list]}

    # -------------------------
    # load
    # -------------------------

    def load_excel(self, path):
        """Loads an Excel file from the given path."""
        wb = load_workbook(path, data_only=True)
        self.sheets = {}

        for ws in wb:
            self.sheets[ws.title] = [
                list(r) for r in ws.iter_rows(values_only=True)
            ]

    def load_csv(self, path):
        """Loads a CSV file from the given path."""
        with open(path, newline="", encoding="utf-8") as f:
            self.sheets = {"Sheet1": list(csv.reader(f))}

    # -------------------------
    # save
    # -------------------------

    def save_excel(self, path):
        """Saves the current sheets to an Excel file at the given path."""
        wb = Workbook()

        first = True
        for name, data in self.sheets.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = name
            first = False

            for row in data:
                ws.append(row)

        wb.save(path)

    def save_csv(self, path, active_sheet_name):
        """Saves the specified sheet to a CSV file at the given path."""
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(self.sheets[active_sheet_name])

    def new_sheet(self, name=None):
        """Creates a new, empty sheet with a unique name."""
        if name is None:
            base = "Sheet"
            i = 1
            while f"{base}{i}" in self.sheets:
                i += 1
            name = f"{base}{i}"

        self.sheets[name] = [[""]]
        return name

    def import_sheet(self, path, sheet_tabs):
        """Imports a sheet from a CSV file and adds it to the SheetsManager."""
        self.load_csv(path)

        # Check if any sheets were loaded
        if not self.sheets:
            return None

        # Use the filename without extension as the sheet name
        sheet_name = os.path.splitext(os.path.basename(path))[0]
        data = self.sheets[
            "Sheet1"
        ]  # csv always load to Sheet1, get the data from Sheet1
        del self.sheets["Sheet1"]  # delete the Sheet1
        self.sheets[sheet_name] = data  # save the sheet with filename

        # Check if the sheet name already exists in the *existing* sheets
        if sheet_name in sheet_tabs.views:
            # Generate a unique name for the imported sheet
            i = 1
            new_name = f"{sheet_name}_{i}"
            while new_name in sheet_tabs.views:
                i += 1
                new_name = f"{sheet_name}_{i}"
            sheet_name = new_name
        return sheet_name, data
